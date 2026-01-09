'''
   these are functions used by the update_data and display_data
   scripts
        
   access these values in other modules by
        import sp500_pe.read_data_func as rd
'''
from openpyxl import load_workbook
import openpyxl.utils.cell as ut_cell
import polars as pl
import polars.selectors as cs

from sp500_earn_price_pkg.helper_func_module \
    import helper_func as hp
from sp500_earn_price_pkg.principal_scripts.code_segments.update_data \
    import write_data_to_files as write
    
import config.config_paths as config
import config.set_params as params

env = config.Fixed_locations
param = params.Update_param

price = param.PRICE_NAME
date = param.DATE_NAME
yr_qtr = param.YR_QTR_NAME
year = param.ANNUAL_DATE
rr_name = env.INPUT_RR_FILE

index = param.IDX_E_COL_NAME
index_type = param.IDX_TYPE_COL_NAME
earnings_type = param.E_TYPE_COL_NAME
earnings_metric = param.E_METRIC_COL_NAME
    
    
def verify_valid_input_files():
    '''
        Verifies that input_dir exists
        Verifies that input_dir contains
            at least one sp input file
            only one TIPS input file
        If true, returns set of sp files 
        if false, returns empty set
    '''
    
    input_dir = env.INPUT_DIR
    sp_glob_str = env.INPUT_SP_FILE_GLOB_STR
    
    if not input_dir.exists():
        hp.message([
            f'{input_dir} does not exist'
        ])
        return set()
    
# both sp and rr files present?
# READ: data files in input_dir
    input_sp_files_set = \
        set(str(f.name) for f in 
            input_dir.glob(sp_glob_str))
    rr_file_lst = [str(f.name) for f in 
                   input_dir.glob(rr_name)]
    
    # if no sp input files
    if not input_sp_files_set:
        hp.message([
            f'no sp input files in \n{input_dir}',
            f'sp file names must conform to {sp_glob_str}'
        ])
        
    # if no unique rr input file
    if not (len(rr_file_lst) == 1):
        hp.message([
            f'no unique {rr_name} in \n{input_dir}',
            f'require one TIPS file with name {rr_name}'
        ])
        # set signal: inputs not valid
        input_sp_files_set = set()
    
    return input_sp_files_set


def ensure_consistent_file_names(names_set):
    '''
    Finds date in each sp input xlsx
        Amends name of input file to:
            'FILE_OUTPUT_WKBK_PREFIX} {DATE_FMT}.xlsx'
        where the date is taken from the
            SHT_EST_NAME sheet, near the top of WKBK_DATE_COL
        amends the names of sp files in input_dir
            to conform to the format above
        sys.exit() if cannot find a date in the file
        Return the set of std names, which includes no redundant
            names: date in file name is date of projection; one
            file per projection date.
        
    https://www.programiz.com/python-programming/datetime/strftime
    https://duckduckgo.com/?q=python+string+to+datetime&t=osx&ia=web
    '''
    
    input_dir = env.INPUT_DIR
    # limit # of rows to read in WKBK_DATE_COL
    max = param.MAX_DATE_ROWS
    date_col = param.WKBK_DATE_COL
    
    output_sp_files_set = set()
    for file in sorted(names_set, reverse= True):
        path_name = input_dir / file
        sheet = find_wk_sheet(
            path_name,
            param.SHT_EST_NAME)
        
        # returns a simple list of vals in date_col
        first_col_list = \
            xlsx_block_reader(sheet, 
                start_row= 1, stop_row= max,
                first_col_ltr= date_col, 
                last_col_ltr= date_col,
                return_simple_list= True)
        
        # find file's date for file's name
        found_date = False
        for item in first_col_list:
            if hp.str_is_date(item, param.DATE_FMT):
                found_date = True
                new_name_file = \
                    f'{env.FILE_OUTPUT_WKBK_PREFIX} {item}.xlsx'
                # rename file, if file is not redundant
                if new_name_file in output_sp_files_set:
                    # delete, more recent version already exists
                    path_name.unlink()
                else:
                    # add name to return set, update dir
                    output_sp_files_set.add(new_name_file)
                    path_name.rename(input_dir / new_name_file)
                break
            
        # traversed first_col_list w/o finding a date
        if not found_date:
            hp.message([
                'ERROR read_data_func.std_names for files',
                'found no date in first {max} rows of',
                f'{input_dir / file}',
                'inspect file for date'
                ])
            write.restore_temp_files()
        
    return output_sp_files_set


def find_wk_sheet(file, sheet_name):
    '''
        Receive 
            path_name, address of xlsx
            sheet_name, name of sheet in xlsx
        Return sheet (openpyxl) for reading
    '''
    path_name = env.INPUT_DIR / file
    try:
        workbook = load_workbook(
                filename= path_name,
                read_only= True, 
                data_only= True)
        return workbook[sheet_name]
    except Exception as e:
        hp.message([
            f'failed to load \n{file}\n{sheet_name}',
            'Check the workbook and sheet, then try again.'
        ])
        write.restore_temp_files()


def xlsx_block_reader(sheet, 
                      start_row= 1, 
                      stop_row= None,
                      first_col_ltr= 'A', 
                      last_col_ltr= None,
                      skip_cols= None, 
                      skip_rows= None,
                      return_simple_list= False,
                      cast_date_to_str= True):
    """
        Returns a rectangular block of data from an xlsx 
            as a list of lists
        From start_row to stop_row, start_col to stop_col
            inclusive
            skip_cols: a list of numbers
            skip_cows: a list of numbers
    """
    if skip_cols is None:
        skip_cols = []
    if skip_rows is None:
        skip_rows = []
        
    # for scanning over rows, for a given col
    if stop_row is None:
        stop_row = sheet.max_row
    
    # for scanning over cols, for a given row
    if last_col_ltr is None:
        last_col_ltr = \
            ut_cell.get_column_letter(sheet.max_column)
    
    rng = sheet[
        f'{first_col_ltr}{start_row}:{last_col_ltr}{stop_row}']
    
    # list of lists
    data = [[cell_.value
             for c_idx, cell_ in enumerate(row)
                 if c_idx not in skip_cols]  
             for r_idx, row in enumerate(rng)
                 if r_idx not in skip_rows]
    
    # option for key finding (not default)
    if return_simple_list:
        # list of items in column, each item is a 1-element list
        # => convert to a simple list of the elements
        # convert any temporal date to date_str
        if first_col_ltr == last_col_ltr:
            return [hp.cast_date_to_str(item[0])
                    for item in data]
    
        # list containing 1 element, a list of items in the row
        # => convert to a simple list by choosing that element
        if start_row == stop_row:
            return data[0]
    
    # otherwise, preserve the list-of-list structure
    # with default option, cast dates in first col to str
    if cast_date_to_str:    
        for row in data:
            row[0] = hp.cast_date_to_str(row[0])
            
    return data
    

def key_finder(cell_list, name,
               start_pos= 0,
               keys_to_find_list= None):
    '''
        Search along the cell_list
        to find the cell that contains a value in
        keys_to_find_list
        
        return the pos of the 1st occurence of
        any item in keys_to_find_list
            (uses 0-based indexing for cell_list)
    '''
    def send_msg():
        hp.message([
            f'In key_finder for {name}',
            f'cell_list: {cell_list}',
            f'start pos: {start_pos}',
            f'keys to find: {keys_to_find_list}'])
        
    if not isinstance(cell_list, list):
        send_msg()
        
    if keys_to_find_list is not None:
        if not isinstance(keys_to_find_list, list):
            send_msg()
    
    for idx, val in enumerate(cell_list[start_pos:]):
        pos = idx + start_pos
        if ((val is None) and
            (keys_to_find_list is None)):
                return pos
        
        if keys_to_find_list is not None:
            if val in keys_to_find_list:
                return pos
    
    # if exhaust the list without finding key
    send_msg()
    

def find_qtrs_without_op_earn(df):
    '''
        Receives pl.df
        Return df with date and yr_qtr
        If df is not empty
            return from col df[date]
            the set of dates
            for which operating eps is null
        otherwise, return empty set
    '''
    df = df.filter(pl.col(param.OP_EPS).is_null())\
           .select(pl.col(date), pl.col(yr_qtr))
    date_set = set(df[date])
    df = df.filter(pl.col(yr_qtr)==min(df[yr_qtr]))\
           .filter(pl.col(date)==min(df[date]))
    return [date_set, df[date].to_list()[0], df[yr_qtr].to_list()[0]]
    
    
def find_yrs_without_rep_earn(df):
    '''
        Receives pl.df
        Return the set of years to update
        If df is not empty
            the set of yrs
            for which rep eps is null
        otherwise, return empty set
    '''
    
    if df.is_empty():
        return pl.DataFrame()
    
    gf = df.filter(
        (pl.col(index_type) == param.SP_IDX_TYPES[0]) &
        (pl.col(earnings_type) == param.EARN_TYPES[1]) &
        (pl.col(earnings_metric) == param.EARN_METRICS[0]) &
        (pl.col(index).is_null()))
    
    return set((item for item in gf[year]))
                

def history_data():
    '''
        Read & temporarily save history from parquet file
        Return as pl.df or empty pl.df
        
        'date' col is pl.String in DATE_FMT
    '''
    if env.OUTPUT_HIST_ADDR.exists():
        with env.OUTPUT_HIST_ADDR.open('rb') as f:
            act_df = pl.read_parquet(f)
        with env.BACKUP_HIST_TEMP_ADDR.open('wb') as f:
            act_df.write_parquet(f)
        return act_df
    else:
        return pl.DataFrame()  

    
def history_loader(file, min_date):
    '''
        Receives an xlsx sheet and dates
            to read from history on the sheet
        Reads two types of data from sheet:
            rectangular block of reported earn
            actual qtr-end prices since last
                qtr that contains reported earn
        Returns a pl.dataframe
            new historical data from the sheet
        sheet is a list of lists,
            a matrix extracted from the sheet
    '''
    
    sheet = find_wk_sheet(file, param.SHT_EST_NAME)
    
    # fetch list of data, from first col of sheet
    # ensure all dates are expressed as strings
    # returns a list of col_vals_list (one for each row),
    # each col_vals_list contains 1 item, [value for col A]
    cell_list = \
        xlsx_block_reader(sheet, 
                    start_row= 1,
                    first_col_ltr= param.WKBK_DATE_COL, 
                    last_col_ltr= param.WKBK_DATE_COL,
                    return_simple_list= True)
    
    # find start pos from values in cell list
    min_pos = key_finder(cell_list, 'hist_loader 1',
                         start_pos = 0,
                         keys_to_find_list= param.HISTORY_KEYS)
    
    # find stop pos: don't reload existing history (if any)
    # but load at least one row of history from current file
    if min_date:
        if min_date <= cell_list[min_pos + 1]:
            stop_key = [min_date]
        else:
            stop_key = [cell_list[min_pos + 1]]
    else:
        stop_key = None
        
    max_pos = key_finder(cell_list, 'hist_loader 2',
                            start_pos= min_pos + 1,
                            keys_to_find_list= stop_key)
    
    # offset to row below start_key, +1 plus
    # offset to adj python indexing to xlsx indexing, +1
    min_row = min_pos + 2
    
    # offset to reconcile xlsx with py 0-based list index: +1
    if stop_key is None:
        # offset to row above the empty row: -1 (1-1=0)
        max_row = max_pos
    else:
        # no additional offset for row found by date key
        max_row = max_pos + 1
    
    df = pl.DataFrame(xlsx_block_reader(sheet, 
                            start_row= min_row, 
                            stop_row= max_row,
                            first_col_ltr= 
                                param.HIST_EPS_COL_START, 
                            last_col_ltr= 
                                param.HIST_EPS_COL_STOP),
                orient= 'row')\
            .select(cs.by_dtype(pl.String, pl.Float64))\
            .cast({cs.float(): pl.Float32})
    df.columns = [name for name in param.HIST_COLUMN_NAMES
                  if not (name == yr_qtr)]
    
    '''
            .with_columns(pl.col(date).map_elements(hp.cast_date_to_str,
                                          return_dtype= pl.String))\
            .with_columns(cs.first().map_elements(hp.cast_date_to_str,
                                          return_dtype= pl.String))\
    '''
    
    recent_prices_df = pl.DataFrame(
            xlsx_block_reader(sheet, 
                            start_row= 
                               min_row - param.PRICE_OFFSET_1, 
                            stop_row= 
                               min_row - param.PRICE_OFFSET_2,
                            first_col_ltr= 
                               param.PRICE_HST_COL_START, 
                            last_col_ltr= 
                               param.PRICE_HST_COL_STOP),
            schema = [date, price],
            orient= 'row')\
        .filter(~pl.col(price).is_null())\
        .cast( {price: pl.Float32})
        
    key_pos_num = key_finder(cell_list, 'history_loader 3',
                             start_pos= 1,
                             keys_to_find_list= param.PRICE_KEYS)
    # offset for py's 0-based indexing
    key_row_num = key_pos_num + 1
    
    [date_, price_] = xlsx_block_reader(sheet, 
                            start_row= key_row_num, 
                            stop_row= key_row_num + 1,
                            first_col_ltr= 
                               param.PRICE_RCNT_COL, 
                            last_col_ltr= 
                               param.PRICE_RCNT_COL,
                            return_simple_list= True)
    
    if date_ not in recent_prices_df[date]:
        recent_prices_df = pl.concat(
                [recent_prices_df, 
                 pl.DataFrame({date: date_,
                        price: price_},
                        orient= 'row')\
                    .cast({price: pl.Float32})],
                how= 'vertical')
        
    df = recent_prices_df.join(df, 
                               on= [price],
                               how= 'full',
                               coalesce=True,)\
            .with_columns(
                pl.when(pl.col(date).is_null())
                       .then(pl.col('date_right'))
                       .otherwise(pl.col(date))
                  .alias(date))\
            .drop(pl.col('date_right'))\
            .sort(by= date)\
            .filter(pl.col(date).is_not_null())\
            .with_columns(pl.col(date).map_batches(
                hp.date_to_year_qtr, 
                return_dtype= pl.String)
            .alias(yr_qtr))
            
    if min_date:
        df = df.filter(pl.col(date) >= min_date)
            
    if (any([item is None
            for item in df[date]])):
        hp.message([
            f'{file} \nmissing date for new entries',
            df[date]
        ])
        write.restore_temp_files()

    return [df, cell_list]


def margin_loader(file, min_yr_qtr, cell_list):
    '''
        read data from s&p excel worksheet
        that contains history for margins
        return df
    '''
    
    sheet = find_wk_sheet(file, param.SHT_EST_NAME)
    
    # find first row of block to be read
    min_pos_num = key_finder(cell_list, 'margin_loader 1',
                             start_pos= 1,
                             keys_to_find_list= param.MARG_KEYS)
    # offset for list's 0-based indexing
    min_row_number = min_pos_num + 1
    
    # find new cell_list along row_pos_num
    cell_list_along_row = \
        xlsx_block_reader(sheet, 
                    start_row= min_row_number, 
                    stop_row= min_row_number,
                    first_col_ltr= param.MARG_KEY_COL,
                    return_simple_list= True)
    
    cell_list_along_row = [item[:4] if isinstance(item, str)
                           else str(item)[:4]
                           for item in cell_list_along_row]
    
    # stop seeking on finding the min yr to update
    # years decrease along the row
    if  min_yr_qtr:
        stop_key = min_yr_qtr[:4]
    else:
        # read to the first blank col (all the years)
        stop_key = param.MARG_STOP_COL_KEY
   
    max_pos_number = \
        key_finder(cell_list_along_row, 'margin_loader 2',
                   start_pos= param.MARG_KEY_POS,
                   keys_to_find_list= [stop_key])
    
    if  min_yr_qtr:
        # offset between 0-based list index and xlsx cols
        max_col_number = max_pos_number + 1
    else:
        # data ends in prev col => offset -1
        # offset between 0-based list and xlsx +1 (1-1=0)
        max_col_number = max_pos_number
    stop_col_ltr = \
        ut_cell.get_column_letter(max_col_number)
    
    # list of lists
    data = xlsx_block_reader(sheet, 
                    start_row= min_row_number + 1, 
                    stop_row= min_row_number + 
                        param.MARG_MAX_ROW_OFFSET,
                    first_col_ltr= 
                        param.MARG_KEY_COL, 
                    last_col_ltr= stop_col_ltr)
    
    # reverse offset for 0-based list vs xlsx cols
    col_names = cell_list_along_row[:max_col_number]
    qtrs = col_names[0]
    
    # build "tall" 2-col DF with 'year_qtr' and 'margin'
    
    df = pl.DataFrame(data, 
                      schema= col_names,
                      orient= 'row')\
                .with_columns(pl.col(qtrs)
                              .map_elements(lambda x: x.split(' ')[0],
                                            return_dtype= str))\
                .cast({cs.float(): pl.Float32})\
                .unpivot(index= qtrs, variable_name= year)
            # index: names of cols to remain cols
            # variable_name: name of col to contain names of cols pivoted
    
    df = df.with_columns(
                pl.struct([qtrs, year])\
                    .map_elements(lambda x: 
                            (f"{x[year]}-{x[qtrs]}"),
                             return_dtype= pl.String)\
                    .alias(yr_qtr))\
            .drop([year, qtrs])\
            .rename({'value': param.MARG_COL_NAME})
            
    return df


def qtrly_loader(file, min_yr_qtr):
    '''
        dates for the actual_df do not nec
        correspond to those loaded here for the
        same quarter, so use yr_qtr
    '''
    sheet = find_wk_sheet(file, param.SHT_QTR_NAME)
        
    # find values down the rows
    cell_list = \
        xlsx_block_reader(sheet, 
                    start_row= 1,
                    stop_row= sheet.max_row,
                    first_col_ltr= param.QTR_HST_COL_START, 
                    last_col_ltr= param.QTR_HST_COL_START,
                    return_simple_list= True)
        
    # convert dates in cell list to yr_qtr
    cell_list_to_yq = [hp.date_to_year_qtr([item])[0] 
                       if hp.str_is_date(item, param.DATE_FMT)
                       else item
                       for item in cell_list]
    
    # stop_key: min date to read
    if min_yr_qtr:
        stop_key = [min_yr_qtr]
    else:
        stop_key = None
        
    min_pos = key_finder(cell_list_to_yq, 'qtrly_loader 1',
                         keys_to_find_list= param.QTR_START_KEYS)
    # offset for 0-based list index: +1
    # offset to move to next row for data: +1
    min_row = min_pos + 2
    
    max_pos = key_finder(cell_list_to_yq, 'qtrly_loader 2',
                         start_pos= min_pos,
                         keys_to_find_list= stop_key)
    # offset for 0-based list index (+1)
    if stop_key is None:
        # offset to row above the empty row: -1 (1-1=0)
        max_row = max_pos
    else:
        # no additional offset for row found by date key
        max_row = max_pos + 1
    
    df = pl.DataFrame(xlsx_block_reader(sheet, 
                            start_row= min_row, 
                            stop_row= max_row,
                            first_col_ltr= 
                                param.QTR_HST_COL_START, 
                            last_col_ltr= 
                                param.QTR_HST_COL_STOP,
                            skip_cols= param.QTR_COL_TO_SKIP),
                      orient= 'row')
    
    '''
             .with_columns(cs.first().map_elements(
                    hp.cast_date_to_str,
                    return_dtype= pl.String))
    '''
    
    df.columns = param.QTR_COLUMN_NAMES
    
    return df.filter(pl.col(date).is_not_null())\
             .with_columns(pl.col(date).map_batches(
                        hp.date_to_year_qtr, 
                        return_dtype= pl.String)
                    .alias(yr_qtr))\
             .cast({cs.float(): pl.Float32})\
             .drop(pl.col(date))
   

def fred_reader(file, min_yr_qtr):
    '''
        read data from FRED excel worksheet
        that contains history for real interest rates
            for dates in dates_set
        return df
    '''
    sheet = find_wk_sheet(file, param.SHT_RR_NAME)
    
    [date, rr] = param.RR_DF_SCHEMA
    
    if min_yr_qtr:
        min_yrqtr = min_yr_qtr
    else:
        min_yrqtr = param.RR_MIN_YR_QTR
    
    # end of qtr data in Fred is dated with the
    # first day of the quarter => convert to yrqtr
    # and drop the dates asap
    '''
        .with_columns(
            pl.col(date).map_elements(
                hp.cast_date_to_str,
                return_dtype= pl.String))
    '''
    return pl.DataFrame(
            xlsx_block_reader(sheet, 
                        start_row= param.RR_MIN_ROW, 
                        stop_row= sheet.max_row,
                        first_col_ltr= param.RR_START_COL, 
                        last_col_ltr= param.RR_STOP_COL),
            schema = param.RR_DF_SCHEMA,
            orient= 'row')\
        .filter(~pl.col(rr).is_null())\
        .cast({rr: pl.Float32})\
        .with_columns(pl.col(date)
                .map_batches(hp.date_to_year_qtr,
                             return_dtype= pl.String)
                .alias(yr_qtr))\
        .filter(pl.col(yr_qtr) >= min_yrqtr)\
        .group_by(yr_qtr)\
        .agg([pl.all().sort_by(date).last()])\
        .sort(by= yr_qtr)\
        .drop(pl.col(date))
        
        
def industry_data():
    '''
    '''
    if env.OUTPUT_IND_ADDR.exists():
        with env.OUTPUT_IND_ADDR.open('rb') as f:
            ind_df = pl.read_parquet(f)
        with env.BACKUP_IND_TEMP_ADDR.open('wb') as f:
            ind_df.write_parquet(f)
        
        col_lst = update_col_names(ind_df.columns)
        ind_df.columns = col_lst
        return ind_df
    
    else:
        return pl.DataFrame()
    
    
def update_col_names(lst):
    '''
    Docstring for update_col_names
    
    :return: Description
    :rtype: Any
    '''
    try:
        idx = lst.index(param.TELECOM_SERV)
        lst[idx] = param.COM_SERV
    except ValueError:
        pass
    return lst
        

def industry_loader(file, years_to_update_set):
    '''
        read data from s&p excel worksheet
        that contains history for industry data
        return df
    '''
    
    sheet = find_wk_sheet(file, param.SHT_IND_NAME)
    
## +++++ read names of industries from 1st col
## find start and stop rows for data block
    cell_list = \
        xlsx_block_reader(sheet, 
                    start_row= 1,
                    stop_row= param.IND_INIT_ROW_ITER,
                    first_col_ltr= param.IND_SRCH_COL, 
                    last_col_ltr= param.IND_SRCH_COL,
                    return_simple_list= True,
                    cast_date_to_str= False)
        
    # find positions of start & stop keys in cell_list
    min_pos = key_finder(cell_list, 'industry_loader 1',
                   start_pos= 0,
                   keys_to_find_list= param.IND_OP_START_KEYS)
    
    max_pos = key_finder(cell_list, 'industry_loader 2',
                   start_pos= min_pos,
                   keys_to_find_list= param.IND_OP_STOP_KEYS)
    
    # convert to row numbers in xlsx
    # offset to adj list indexing to xlsx indexing + 1
    start_row = min_pos + 1
    # additional offset, - 1, because stop key is one row below data block
    stop_row = max_pos
    
    # remove parentheticals and 'SP 500' from ind names
    ind_name = [' '.join(item.split(" (")[0].split(' ')[2:])
           for item in cell_list[min_pos:max_pos]]
    
    ind_name = update_col_names(ind_name)
    # set first name
    ind_name[0] = param.IDX_E_COL_NAME
    
## +++++  find data ids ++++++++++++++++++++
## find first and last col letters for block data
    # search first row of dates, 2 rows above start row (offset -2)
    # for years that are not in years_no_update_set
    # get data from row with years in the cols
    
    cell_list_from_cols = \
        xlsx_block_reader(sheet, 
                    start_row= start_row + param.IND_OFFSET,
                    stop_row= start_row + param.IND_OFFSET,
                    first_col_ltr= param.IND_DATA_START_COL,
                    return_simple_list= True,
                    cast_date_to_str= False)
        
    min_pos = key_finder(cell_list_from_cols, 'industry loader 3',
                   start_pos= 0,
                   keys_to_find_list= param.IND_DATA_FIRST_COL_KEY)
    
    max_pos = key_finder(cell_list_from_cols, 'industry loader 4',
                   start_pos= min_pos,
                   keys_to_find_list= param.IND_DATA_LAST_COL_KEY)
    
    # find new min_pos: year > max(years_no_update_set)
    if years_to_update_set:
        # trim cell_list_from_cols to new start col
        min_year_to_update = int(min(years_to_update_set))
        while min_pos <= max_pos - 2:  
            if int(cell_list_from_cols[min_pos][:4]) >= \
               min_year_to_update:
                break
            min_pos += 2 #for each year, EPS col and PE col
            
    # convert to col numbers in xlsx
    # offset to adj list indexing to xlsx indexing + 1
    min_pos_xlsx_col = \
          min_pos + param.IND_DATA_START_COL_OFFSET + 1
    # additional offset, - 1, stop key is one col after data block
    max_pos_xlsx_col = \
         max_pos + param.IND_DATA_START_COL_OFFSET
        
    first_col_ltr = ut_cell.get_column_letter(min_pos_xlsx_col)
    last_col_ltr = ut_cell.get_column_letter(max_pos_xlsx_col)
    
    # create names of columns for finished df
    cols_schema =  cell_list_from_cols[min_pos : max_pos]
    
    cols_e = [col for col in cols_schema
              if param.EPS_MK in col]
    cols_pe = [col for col in cols_schema
               if param.PE_MK in col]
    years = [item[:4] for item in cols_e]
    
## loop over the 4 SP indexes and the 2 types of earnings
    first_base_row = start_row
    end_base_row = stop_row
    first_pass_through_loops = True
    
    for index_type in param.SP_IDX_TYPES:
        for e_type in param.EARN_TYPES:
            if e_type == param.EARN_TYPES[0]:
                first_row = first_base_row
                end_row = end_base_row
            else:
                first_row += param.OP_REP_OFFSET
                end_row += param.OP_REP_OFFSET
            
            gf = separate_transpose_combine_df(sheet, 
                first_row, end_row, first_col_ltr, last_col_ltr,
                ind_name, cols_schema, cols_e, cols_pe, years,
                index_type, e_type)
            
            # join the 8 dfs, as they are produced
            if first_pass_through_loops:
                df = gf
                first_pass_through_loops = False
            else:
                df = pl.concat([df, gf],
                               how= 'vertical')
        # do at end of all index_type loops, except the last
        if not (index_type == param.SP_IDX_TYPES[-1:]):
            first_base_row += param.IDX_OFFSET
            end_base_row += param.IDX_OFFSET
    
    return df


def separate_transpose_combine_df(sheet, 
        start_row, stop_row, first_col_ltr, last_col_ltr,
        ind_name, cols_schema, cols_e, cols_pe, years,
        index_type, e_type):
    '''
        Read block of annual data from sheet
        Separate level of earnings from pe ratios
        Rejoin these two mini blocks via vertical concat
        Transpose block of data from ind x year to
            yr x ind
        Add 3 columns to identify the data in each row
            SP index {same for all rows in a block}, 
            type of earnings (op or rep) {ditto},
            level or pe ratio {same for all rows in each mini block}
    '''
    
    data = xlsx_block_reader(sheet, 
                             start_row= start_row,
                             stop_row= stop_row,
                             first_col_ltr= first_col_ltr,
                             last_col_ltr= last_col_ltr,
                             return_simple_list= False,
                             cast_date_to_str= False)

    df = pl.DataFrame(data, 
                      schema= cols_schema,
                      orient= 'row')\
           .cast({cs.float(): pl.Float32})
           
    # select and pivot to years for rows, data type for cols
    df_pe = hp.transpose_df(df, ind_name, param.EARN_METRICS[1],
                            index_type, e_type, cols_pe, years)
  
    df_e  = hp.transpose_df(df, ind_name, param.EARN_METRICS[0], 
                            index_type, e_type, cols_e, years)
    
    return pl.concat([df_e, df_pe],
                     how= 'vertical')
    
    
def projection_data():
    '''
        Read & create temp backup for existing 
        
        Proj_dict contains projections
            key: yr_qtr, in which the proj was made
            val: df containing the projs for future dates
        The projs for each yr_qtr are the proj for the 
            latest date in the quarter (.xlsx workbooks)
            
        Return proj_dict, contains keys for all yr_qtrs
            to date (data begins in 2017)
    '''
    if env.OUTPUT_PROJ_ADDR.exists():
        with env.OUTPUT_PROJ_ADDR.open('rb') as f:
            proj_hist_df = pl.read_parquet(f)
        with env.BACKUP_PROJ_TEMP_ADDR.open('wb') as f:
            proj_hist_df.write_parquet(f)
        proj_dict = proj_hist_df.to_dict(as_series= False)
    else:
        proj_dict = dict()
    
    return proj_dict


def proj_loader(file):
    '''
    '''
    hp.message([
        f'reading projections from',
        f'input file: {file}'])

    # takes an iterable, returns a list
    year_quarter = hp.date_to_year_qtr([
                        file.split('.')[0][-10:]
                   ])[0]
    
    sheet = find_wk_sheet(file, param.SHT_EST_NAME)
    
    # find first row of block of proj to read
    cell_list = \
        xlsx_block_reader(sheet, 
                    start_row= 1,
                    first_col_ltr= param.WKBK_DATE_COL, 
                    last_col_ltr= param.WKBK_DATE_COL,
                    return_simple_list= True)\
                        [:param.PROJ_MAX_LIST]
        
    # find positions of start & stop keys in cell_list
    min_pos = key_finder(cell_list,'proj loader 1',
                   start_pos= 0,
                   keys_to_find_list= \
                       param.PROJ_ROW_START_KEYS)
    max_pos = key_finder(cell_list, 'proj loader 2',
                         start_pos= min_pos,
                         keys_to_find_list= \
                             param.PROJ_ROW_STOP_KEYS)

    # offset to row below start_key, +1
    # offset to adj python indexing to xlsx indexing, +1
    min_row = min_pos + 2
    max_row = max_pos
    
    df = pl.DataFrame(xlsx_block_reader(sheet, 
                         start_row= min_row, 
                         stop_row= max_row,
                         first_col_ltr= 
                                param.HIST_EPS_COL_START, 
                         last_col_ltr= 
                                param.HIST_EPS_COL_STOP,
                         skip_cols= param.PROJ_COLS_TO_SKIP),
                orient= 'row')\
            .select(cs.by_dtype(pl.String, pl.Float64))\
            .cast({cs.float(): pl.Float32})
            
    df.columns = [name for name in param.PROJ_COLUMN_NAMES
                  if not (name == yr_qtr)]
    
    df = df.with_columns(pl.col(date).map_batches(
                hp.date_to_year_qtr, 
                return_dtype= pl.String)
            .alias(yr_qtr))
    
    return [df, year_quarter]