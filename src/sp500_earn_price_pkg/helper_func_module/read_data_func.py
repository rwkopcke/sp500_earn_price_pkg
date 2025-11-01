'''
   these are functions used by the update_data and display_data
   scripts
        
   access these values in other modules by
        import sp500_pe.read_data_func as rd
'''
from datetime import datetime


from openpyxl import load_workbook
import openpyxl.utils.cell as ut_cell
from openpyxl.utils import coordinate_to_tuple
import polars as pl
import polars.selectors as cs

from sp500_earn_price_pkg.helper_func_module \
    import helper_func as hp
    
import sp500_earn_price_pkg.config.config_paths as config
import sp500_earn_price_pkg.config.set_params as params

env = config.Fixed_locations()
rd_param = params.Update_param()
    
    
def verify_valid_input_files():
    '''
        Verifies that input_dir exists
        Verifies that input_dir contains
            at least one sp input file
            only one TIPS input file
        If true, returns set of sp files 
        if false, returns empty set
    '''
    
    input_dir = env.INPUT_DIR
    sp_glob_str = env.INPUT_SP_FILE_GLOB_STR
    rr_name = env.INPUT_RR_FILE
    
    if not input_dir.exists():
        hp.message([
            f'{input_dir} does not exist'
        ])
        return {}

# both sp and rr files present?
# READ: data files in input_dir
    input_sp_files_set = \
        set(str(f.name) for f in 
            input_dir.glob(sp_glob_str))
    rr_file_lst = [str(f.name) for f in 
                   input_dir.glob(rr_name)]
    
    # if no sp input files
    if not input_sp_files_set:
        hp.message([
            f'no sp input files in \n{input_dir}',
            f'sp file names must conform to {sp_glob_str}'
        ])
        
    # if no unique rr input file
    if not (len(rr_file_lst) == 1):
        hp.message([
            f'no unique {rr_name} in \n{input_dir}',
            f'require one TIPS file with name {rr_name}'
        ])
        # set signal: inputs not valid
        input_sp_files_set = {}
    
    return input_sp_files_set


def ensure_consistent_file_names(names_set):
    '''
    Finds date in each sp input xlsx
        Amends name of input file to:
            'FILE_OUTPUT_WKBK_PREFIX} {DATE_FMT}.xlsx'
        where the date is taken from the
            SHT_EST_NAME sheet, near the top of WKBK_DATE_COL
        amends the names of sp files in input_dir
            to conform to the format above
        quit() if cannot find a date in the file
        Returns the set of std names
        
    https://www.programiz.com/python-programming/datetime/strftime
    https://duckduckgo.com/?q=python+string+to+datetime&t=osx&ia=web
    '''
    
    input_dir = env.INPUT_DIR
    # limit # of rows to read in WKBK_DATE_COL
    max = rd_param.MAX_DATE_ROWS
    date_col = rd_param.WKBK_DATE_COL
    
    output_sp_files_set = set()
    for file in names_set:
        path_name = input_dir / file
        sheet = find_wk_sheet(
            path_name,
            rd_param.SHT_EST_NAME)
        
        # returns a simple list of vals in date_col
        first_col_list = \
            xlsx_block_reader(sheet, 
                start_row= 1, stop_row= max,
                first_col_ltr= date_col, 
                last_col_ltr= date_col,
                return_simple_list= True)
        
        # find file's date for file's name
        for item in first_col_list:
            if hp.str_is_date(item, rd_param.DATE_FMT):
                new_name_file = \
                    f'{env.FILE_OUTPUT_WKBK_PREFIX} {item}.xlsx'
                # add name to return set
                output_sp_files_set.add(new_name_file)
                # rename file
                path_name.rename(input_dir / new_name_file)
                break
        
        # full inspection of col_date_list w/o finding a date
        if not item:
            hp.message([
                'ERROR read_data_func.std_names for files',
                'found no date in first {max} rows of',
                f'{input_dir / file}',
                'inspect file for date'
                ])
            quit()
        
    return output_sp_files_set


def find_wk_sheet(file, sheet_name):
    '''
        Receive 
            path_name, address of xlsx
            sheet_name, name of sheet in xlsx
        Return sheet (openpyxl) for reading
    '''
    path_name = env.INPUT_DIR / file
    workbook = load_workbook(
            filename= path_name,
            read_only= True, 
            data_only= True)
    return workbook[sheet_name]


def xlsx_block_reader(sheet, 
                      start_row= 1, 
                      stop_row= None,
                      first_col_ltr= 'A', 
                      last_col_ltr= None,
                      skip_cols= None, 
                      skip_rows= None,
                      return_simple_list= False,
                      cast_date_to_str= True):
    """
        Returns a rectangular block of data from an xlsx 
            as a list of lists
        From start_row to stop_row, start_col to stop_col
            inclusive
            skip_cols: a list of numbers
            skip_cows: a list of numbers
    """
    if skip_cols is None:
        skip_cols = []
    if skip_rows is None:
        skip_rows = []
        
    # for scanning over rows, for a given col
    if stop_row is None:
        stop_row = sheet.max_row
    
    # for scanning over cols, for a given row
    if last_col_ltr is None:
        last_col_ltr = \
            ut_cell.get_column_letter(sheet.max_column)
    
    rng = sheet[
        f'{first_col_ltr}{start_row}:{last_col_ltr}{stop_row}']
    
    # list of lists
    data = [[cell_.value
             for c_idx, cell_ in enumerate(row)
                 if c_idx not in skip_cols]  
             for r_idx, row in enumerate(rng)
                 if r_idx not in skip_rows]
    
    # option for key finding (not default)
    if return_simple_list:
        # list of items in column, each item is a 1-element list
        # => convert to a simple list of the elements
        # convert any temporal date to date_str
        if first_col_ltr == last_col_ltr:
            return [hp.cast_date_to_str(item[0])
                    for item in data]
    
        # list containing 1 element, a list of items in the row
        # => convert to a simple list by choosing that element
        if start_row == stop_row:
            return data[0]
    
    # otherwise, preserve the list-of-list structure
    # with default option, cast dates in first col to str
    if cast_date_to_str:    
        for row in data:
            row[0] = hp.cast_date_to_str(row[0])
            
    return data
    

def find_keys_in_list(cell_list, start_pos, keys_to_find):
    '''
        This function returns the first position of an item
        in keys_to_find in the cell_list.
        
        Search starts at start_pos
    '''
    
    for idx, val in enumerate(cell_list[start_pos:]):
        if val is None:
            if keys_to_find is None:
                return idx + start_pos
        else:
            if keys_to_find is not None:
                if val in keys_to_find:
                    return idx + start_pos
    
    # if exhaust the list without finding key
    hp.message([
        f'In find_keys_in_list',
        f'cell_list: {cell_list}',
        f'start pos: {start_pos}',
        f'keys to find: {keys_to_find}'
    ])
    quit() 
        
def key_finder(cell_list= None,
               start_keys= ['no start key'],
               stop_keys= ['no stop key']):
    '''
        Search along the cell_list
        to find the cell that contains a value in
        [start_keys], then the cell that contains
        a value in [stop_keys]
        
        if the call specifies no start or stop keys
        then this routine skips that search
        
        return the pos of the start and stop keys
        in the list, 0-based indexing
    '''
    def send_error():
        hp.message([
            'ERROR In read_data_func.py, key_finder()',
            'sone of the inputs are not valid',
            f'keys must be non-null list: \n{cell_list}',
            f'start_keys must be [str] or None: {start_keys}',
            f'stop_keys must be [str] or None: {stop_keys}'])
        quit()
        
    if ((cell_list is None) or
        (not cell_list) or
        (not isinstance(cell_list, list))):
        send_error()
        
    # start & stop keys must be lists of str or None
    # then keys must be a list of str or []
    if ((not isinstance(start_keys, list)) or
        (not start_keys) or
        (not all(isinstance(item, str) for item in start_keys))):
        if (start_keys is not None):
            send_error()
        
    if ((not isinstance(stop_keys, list)) or
        (not stop_keys) or
        (not all(isinstance(item, str) for item in stop_keys))):
        if stop_keys is not None:
            send_error()
     
    if start_keys == ['no start key']:
        min_pos = None
    else:
        min_pos = find_keys_in_list(cell_list,
                                    start_pos= 0,
                                    keys_to_find= start_keys)
    
    if stop_keys == ['no stop key']:
        max_pos = None
    else:
        if min_pos is None:
            start_pos = 0
        else:
            start_pos = min_pos
            max_pos = find_keys_in_list(cell_list,
                                    start_pos= start_pos,
                                    keys_to_find= stop_keys)
    return [min_pos, max_pos]


def read_existing_history():
    '''
        Read saved history from parquet file, and
        Return as pl.df or empty pl.df
        
        Ensures 'date' col in pl.String in DATE_FMT
    '''
    if env.OUTPUT_HIST_ADDR.exists():
        return pl.read_parquet(env.OUTPUT_HIST_ADDR)\
                 .cast({'date': pl.String})
    else:
        return pl.DataFrame()
    

def find_qtrs_without_op_earn(df):
    '''
        Receives pl.df
        Return the set of rows to update
        If df is not empty
            return from col df[yr_qtrs]
            the set of yr_qtrs
            for which operating eps is null
        otherwise, return empty set
    '''
    
    if not df.is_empty():
        return set(
            pl.Series(df.filter(pl.col('op_eps').is_null())
                        .select(pl.col('date')))
                        .to_list())
    else:
        return set()  

    
def update_history(file, dates_set):
    '''
        Receives an xlsx sheet and dates
            to read from history on the sheet
        Reads two types of data from sheet:
            rectangular block of reported earn
            actual qtr-end prices since last
                qtr that contains reported earn
        Returns a pl.dataframe
            new historical data from the sheet
        sheet is a list of lists,
            a matrix extracted from the sheet
    '''
    sheet = find_wk_sheet(file, rd_param.SHT_EST_NAME)
    price = rd_param.PRICE_NAME
    date = rd_param.DATE_NAME
    
    # fetch list of data, from first col of sheet
    # ensure all dates are expressed as strings
    # returns a list of col_vals_list (one for each row),
    # each col_vals_list contains 1 item, [value for col A]
    cell_list = \
        xlsx_block_reader(sheet, 
                    start_row= 1,
                    first_col_ltr= rd_param.WKBK_DATE_COL, 
                    last_col_ltr= rd_param.WKBK_DATE_COL,
                    return_simple_list= True)

    # stop_key: min date to read, a list of 1 str
    if dates_set:
        stop_key = [min(dates_set)]
    else:
        stop_key = None
        
    # find positions of start & stop keys in cell_list
    [min_pos, max_pos] = \
        key_finder(cell_list,
                   start_keys= rd_param.HISTORY_KEYS,
                   stop_keys= stop_key)
    # offset to row below start_key, +1
    # offset to adj python indexing to xlsx indexing, +q
    min_row = min_pos + 2
    max_row = max_pos + 1
    
    df = pl.DataFrame(xlsx_block_reader(sheet, 
                         start_row= min_row, 
                         stop_row= max_row,
                         first_col_ltr= 
                                rd_param.HIST_EPS_COL_START, 
                         last_col_ltr= 
                                rd_param.HIST_EPS_COL_STOP),
                orient= 'row')\
            .select(cs.by_dtype(pl.String, pl.Float64))\
            .cast({cs.float(): pl.Float32})
    df.columns = rd_param.HIST_COLUMN_NAMES
    
    '''
            .with_columns(pl.col(date).map_elements(hp.cast_date_to_str,
                                          return_dtype= pl.String))\
            .with_columns(cs.first().map_elements(hp.cast_date_to_str,
                                          return_dtype= pl.String))\
    '''
    
    recent_prices_df = pl.DataFrame(
            xlsx_block_reader(sheet, 
                            start_row= 
                               min_row - rd_param.PRICE_OFFSET_1, 
                            stop_row= 
                               min_row - rd_param.PRICE_OFFSET_2,
                            first_col_ltr= 
                               rd_param.PRICE_HST_COL_START, 
                            last_col_ltr= 
                               rd_param.PRICE_HST_COL_STOP),
            schema = [date, price],
            orient= 'row')\
        .filter(~pl.col(price).is_null())\
        .cast( {price: pl.Float32})
        
    key_pos_num = find_keys_in_list(cell_list,
                            start_pos= 1,
                            keys_to_find= rd_param.PRICE_KEYS)
    # offset for py's 0-based indexing
    key_row_num = key_pos_num + 1
    
    [date_, price_] = xlsx_block_reader(sheet, 
                            start_row= key_row_num, 
                            stop_row= key_row_num + 1,
                            first_col_ltr= 
                               rd_param.PRICE_RCNT_COL, 
                            last_col_ltr= 
                               rd_param.PRICE_RCNT_COL,
                            return_simple_list= True)
    current_price_df = \
        pl.DataFrame({date: date_,
                      price: price_},
                     orient= 'row')\
          .cast({price: pl.Float32})
    
    recent_prices_df = pl.concat(
        [recent_prices_df, current_price_df],
        how= 'vertical')
        
    df = recent_prices_df.join(df, 
                               on= [price],
                               how= 'full',
                               coalesce=True,)\
            .with_columns(
                pl.when(pl.col(date).is_null())
                       .then(pl.col('date_right'))
                       .otherwise(pl.col(date))
                  .alias(date))\
            .drop(pl.col('date_right'))\
            .sort(by= date)\
            .with_columns(pl.col(date).map_batches(
                hp.date_to_year_qtr, 
                return_dtype= pl.String)
            .alias(rd_param.YR_QTR_NAME))
            
    if (any([item is None
            for item in df[date]])):
        hp.message([
            f'{file} \nmissing date for new entries',
            df[date]
        ])
        quit()

    return [df, cell_list]


def margin_loader(file, dates_set, cell_list):
    '''
        read data from s&p excel worksheet
        that contains history for margins
        return df
    '''
    
    sheet = find_wk_sheet(file, rd_param.SHT_EST_NAME)
    
    # find first row of block to be read
    row_pos_num = find_keys_in_list(cell_list,
                            start_pos= 1,
                            keys_to_find= rd_param.MARG_KEYS)
    # offset for py's 0-based indexing
    row_pos_num += 1
    
    # find new cell_list along row_pos_num
    cell_list_along_row = \
        xlsx_block_reader(sheet, 
                    start_row= row_pos_num, 
                    stop_row= row_pos_num,
                    first_col_ltr= rd_param.MARG_KEY_COL,
                    return_simple_list= True)
    # because first item in cell_list_along_row is the sequence
    # col values across the row\
   
    [_, max_col] =\
        key_finder(cell_list_along_row,
                   start_keys= rd_param.MARG_KEYS,
                   stop_keys= rd_param.MARG_STOP_COL_KEY)
    
    stop_col_ltr = ut_cell.get_column_letter(max_col)
    
    # list of lists
    data = xlsx_block_reader(sheet, 
                    start_row= row_pos_num, 
                    stop_row= row_pos_num + rd_param.MARG_MAX_ROW_OFFSET,
                    first_col_ltr= 
                        rd_param.MARG_KEY_COL, 
                    last_col_ltr= stop_col_ltr)
     
    # data_values omits the first row (col headers) from data
    data_values = [row for row in data[1:]]
    
    # omit the * for 2008, take first entry (yr) in list data[0]
    col_names = [str(item).split('*')[0] for item in data[0]]
    qtrs = col_names[0]
    
    # build "tall" 2-col DF with 'year_qtr' and 'margin'
    df = pl.DataFrame(data_values, 
                      schema= col_names,
                      orient= 'row')\
                .with_columns(pl.col(qtrs)
                              .map_elements(lambda x: x.split(' ')[0],
                                            return_dtype= str))\
                .cast({cs.float(): pl.Float32})\
                .unpivot(index= qtrs, variable_name= rd_param.ANNUAL_DATE)
            # index: names of cols to remain cols
            # variable_name: name of col to contain names of cols pivoted
    
    df = df.with_columns(
                pl.struct([qtrs, rd_param.ANNUAL_DATE])\
                    .map_elements(lambda x: 
                            (f"{x[rd_param.ANNUAL_DATE]}-{x[qtrs]}"),
                             return_dtype= pl.String)\
                    .alias(rd_param.YR_QTR_NAME))\
            .drop([rd_param.ANNUAL_DATE, qtrs])\
            .rename({'value': 'op_margin'})
            
    return df

def qtrly_loader(file, dates_set):
    '''
    '''
    sheet = find_wk_sheet(file, rd_param.SHT_QTR_NAME)
    
    # find values down the rows
    cell_list = \
        xlsx_block_reader(sheet, 
                    start_row= 1,
                    stop_row= rd_param.QTR_INIT_ROW_ITER,
                    first_col_ltr= rd_param.QTR_HST_COL_START, 
                    last_col_ltr= rd_param.QTR_HST_COL_START,
                    return_simple_list= True)
    
    # stop_key: min date to read
    if dates_set:
        stop_key = [min(dates_set)]
    else:
        stop_key = None

    [min_row, max_row] = key_finder(cell_list,
                                    start_keys= rd_param.QTR_START_KEYS,
                                    stop_keys= stop_key)
    # offset 1 for 0-based list index; 1 for row below key
    min_row += 2
    # max_row offset is 1
    max_row += 1
    
    df = pl.DataFrame(xlsx_block_reader(sheet, 
                            start_row= min_row, 
                            stop_row= max_row,
                            first_col_ltr= 
                                rd_param.QTR_HST_COL_START, 
                            last_col_ltr= 
                                rd_param.QTR_HST_COL_STOP,
                            skip_cols= rd_param.QTR_COL_TO_SKIP),
                      orient= 'row')
    
    '''
             .with_columns(cs.first().map_elements(
                    hp.cast_date_to_str,
                    return_dtype= pl.String))
    '''
    
    df.columns = rd_param.QTR_COLUMN_NAMES
    
    return df.with_columns(pl.col(rd_param.DATE_NAME).map_batches(
                        hp.date_to_year_qtr, 
                        return_dtype= pl.String)
                    .alias(rd_param.YR_QTR_NAME))\
             .cast({cs.float(): pl.Float32})\
             .drop(pl.col(rd_param.DATE_NAME))
   

def fred_reader(file, dates_set):
    '''
        read data from FRED excel worksheet
        that contains history for real interest rates
            for dates in dates_set
        return df
    '''
    sheet = find_wk_sheet(file, rd_param.SHT_RR_NAME)
    
    [date, rr] = rd_param.RR_DF_SCHEMA
    yr_qtr = rd_param.YR_QTR_NAME
    min_yrqtr = hp.date_to_year_qtr(
        [hp.cast_date_to_str(min(dates_set))])
    
    # end of qtr data in Fred is dated with the
    # first day of the quarter => convert to yrqtr
    # and drop the dates asap
    '''
        .with_columns(
            pl.col(date).map_elements(
                hp.cast_date_to_str,
                return_dtype= pl.String))
    '''
    return pl.DataFrame(
            xlsx_block_reader(sheet, 
                        start_row= rd_param.RR_MIN_ROW, 
                        stop_row= sheet.max_row,
                        first_col_ltr= rd_param.RR_START_COL, 
                        last_col_ltr= rd_param.RR_STOP_COL),
            schema = rd_param.RR_DF_SCHEMA,
            orient= 'row')\
        .filter(~pl.col(rr).is_null())\
        .cast({rr: pl.Float32})\
        .with_columns(pl.col(date)
                .map_batches(hp.date_to_year_qtr,
                             return_dtype= pl.String)
                .alias(yr_qtr))\
        .filter(pl.col(yr_qtr) >= min_yrqtr)\
        .group_by(yr_qtr)\
        .agg([pl.all().sort_by(date).last()])\
        .sort(by= yr_qtr)\
        .drop(pl.col(date))
        

def industry_loader(file, years_no_update_set):
    '''
        read data from s&p excel worksheet
        that contains history for industry data
        return df
    ''' 
    sheet = find_wk_sheet(file, rd_param.SHT_IND_NAME)
    
## +++++ read names of industries from 1st col
## find start and stop rows for data block
    cell_list = \
        xlsx_block_reader(sheet, 
                    start_row= 1,
                    stop_row= rd_param.IND_INIT_ROW_ITER,
                    first_col_ltr= rd_param.IND_SRCH_COL, 
                    last_col_ltr= rd_param.IND_SRCH_COL,
                    return_simple_list= True,
                    cast_date_to_str= False)
        
    # find positions of start & stop keys in cell_list
    [min_pos, max_pos] = \
        key_finder(cell_list,
                   start_keys= rd_param.IND_OP_START_KEYS,
                   stop_keys= rd_param.IND_OP_STOP_KEYS)
        
    # offset to row below start_key, +1
    # offset to adj python indexing to xlsx indexing,
    start_row = min_pos + 1
    stop_row = max_pos
    
    # remove parentheticals
    ind = [item.split(' (')[0]
           for item in cell_list[min_pos:max_pos]]

    # remove 'S&P 500' and replace space with '_'
    ind_name = ['_'.join(item.rstrip().split(' ')[2:])
                for item in ind]
    # set first name
    ind_name[0] = rd_param.FIRST_IND_NAME
    
## +++++  find data ids ++++++++++++++++++++
## find first and last col letters for block data
    # search the first row of dates (offset -2)
    # for years that are not in years_no_update_set
    # get data from row with years in the cols
    
    cell_list_from_cols = \
        xlsx_block_reader(sheet, 
                    start_row= start_row + rd_param.IND_OFFSET,
                    stop_row= start_row + rd_param.IND_OFFSET,
                    first_col_ltr= rd_param.IND_DATA_START_COL,
                    return_simple_list= True,
                    cast_date_to_str= False)

    [min_pos, max_pos] = \
        key_finder(cell_list_from_cols,
                   start_keys= rd_param.IND_DATA_FIRST_COL_KEY,
                   stop_keys= rd_param.IND_DATA_LAST_COL_KEY)
        
    # search list for first year that is not in years_no_update_set
    max_year_no_update = max(years_no_update_set)

    min_pos_2 = min_pos - 1
    while min_pos_2 <= max_pos:
        min_pos_2 += 1
        if cell_list_from_cols[min_pos_2][:4] > max_year_no_update:
            break
    
    # includes correction for list 0-based indexing
    min_pos_xlsx_col = \
          min_pos_2 + rd_param.IND_DATA_START_COL_OFFSET + 1
    max_pos_xlsx_col = \
         max_pos + rd_param.IND_DATA_START_COL_OFFSET
        
    first_col_ltr = ut_cell.get_column_letter(min_pos_xlsx_col)
    last_col_ltr = ut_cell.get_column_letter(max_pos_xlsx_col)
    
    # get names for data
    date_data_names = [f'{name[:4]} {name[-3:]}'
                       for name in 
                       cell_list_from_cols[min_pos_2:max_pos]]
    columns_pe = [col
                  for col in date_data_names
                  if 'P/E' in col]
    columns_e = [col 
                 for col in date_data_names
                 if 'EPS' in col]
    years = [item[0:4] for item in columns_e]
    
## ++++++++++++ op data +++++++++++++++++
    # fetch op e by industry, including row with headings
    # list of lists for each row
    earn_type = rd_param.OP_EPS
    ratio_type = rd_param.OP_PE
    
    df = separate_pivot_combine_eps_pe_df(sheet, 
            start_row, stop_row, first_col_ltr, last_col_ltr,
            ind_name, date_data_names, columns_pe, columns_e, 
            years, earn_type, ratio_type)
    
## ++++++++++++ rep data +++++++++++++++++
    start_row += rd_param.IND_REP_OFFSET_KEYS
    stop_row = stop_row + rd_param.IND_REP_OFFSET_KEYS
    earn_type = rd_param.REP_EPS
    ratio_type = rd_param.REP_PE
    
    gf = separate_pivot_combine_eps_pe_df(sheet, 
            start_row, stop_row, first_col_ltr, last_col_ltr,
            ind_name, date_data_names, columns_pe, columns_e, 
            years, earn_type, ratio_type)

## ++++++++++++ Combine OP and REP +++++++
    df = df.join(gf,
                   on= rd_param.ANNUAL_DATE,
                   how= 'left',
                   coalesce= True)
    return df


def separate_pivot_combine_eps_pe_df(sheet, 
        start_row, stop_row, first_col_ltr, last_col_ltr,
        ind_name, date_data_names, columns_pe, columns_e, 
        years, earn_type, ratio_type):
    '''
    '''
    data = xlsx_block_reader(sheet, 
                             start_row= start_row,
                             stop_row= stop_row,
                             first_col_ltr= first_col_ltr,
                             last_col_ltr= last_col_ltr,
                             return_simple_list= False,
                             cast_date_to_str= False)

    df = pl.DataFrame(data, 
                      schema= date_data_names,
                      orient= 'row')\
           .cast({cs.float(): pl.Float32})
    
    # select and pivot to years for rows, data type for cols
    df_pe = hp.gen_sub_df(df, ind_name, ratio_type, 
                          columns_pe, years)
    df_e  = hp.gen_sub_df(df, ind_name, earn_type, 
                          columns_e, years)
    
    return df_e.join(df_pe,
                   on= rd_param.ANNUAL_DATE,
                   how= 'left',
                   coalesce= True)
